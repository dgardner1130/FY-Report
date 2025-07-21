import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
import tempfile
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.formatting.rule import CellIsRule
import asana
from asana.rest import ApiException
from pprint import pprint
from datetime import datetime

section = st.sidebar.radio("Select Report Type", ["📊 FY Review Report", "🏗️ MDP Annual Report"])

if section == "📊 FY Review Report":
    # UI
    st.title("📊 FY Project Report Creator")
    st.write("Upload a CSV file and specify the fiscal year start.")

    uploaded_file = st.file_uploader("Upload Fiscal Year Data", type="csv")
    fiscal_year_start_month = 7
    fiscal_year_start_year = st.number_input("Fiscal Year Start Year", value=2024, step=1)

    if uploaded_file:
        data = pd.read_csv(uploaded_file)
        data['Date Comment Letter Sent'] = pd.to_datetime(data['Date Comment Letter Sent'], errors='coerce')
        data['Date Submitted'] = pd.to_datetime(data['Date Submitted'], errors='coerce')

        fiscal_months = list(range(fiscal_year_start_month, 13)) + list(range(1, fiscal_year_start_month))
        months = {}

        for month in fiscal_months:
            year = fiscal_year_start_year if month >= fiscal_year_start_month else fiscal_year_start_year + 1
            start_date = pd.Timestamp(year, month, 1)
            end_date = pd.Timestamp(year+1, 1, 1) - pd.Timedelta(days=1) if month == 12 else pd.Timestamp(year, month+1, 1) - pd.Timedelta(days=1)

            month_df = data[(data['Date Comment Letter Sent'] >= start_date) & 
                            (data['Date Comment Letter Sent'] <= end_date)][
                                ['Date Submitted', 'Development Name', 'Project No', 
                                 'Review Cycle - ENG', 'Review Cycle - SUR', 'Review Cycle - PLN',
                                 'Date Comment Letter Sent']
                            ].copy()
    
            month_df['Length of Review'] = (month_df['Date Comment Letter Sent'] - month_df['Date Submitted']).dt.days
            key = f'{start_date.strftime("%b %Y")}'
            months[key] = month_df

        reviewValues = [df['Length of Review'].tolist() for df in months.values()]
        lenResults = [len(lst) for lst in reviewValues]
        countExceeds30 = [sum(1 for x in lst if x > 30) for lst in reviewValues]
        shortCount = [total - long for total, long in zip(lenResults, countExceeds30)]
        avgLength = [round(np.mean(lst), 2) if lst else 0 for lst in reviewValues]
        percentage = [round((total - long)/total, 3) if total else 0 for total, long in zip(lenResults, countExceeds30)]

        summary_df = pd.DataFrame({
            'Month': list(months.keys()),
            'Total Reviews': lenResults,
            '> 30 Days': countExceeds30,
            'Average Review Length': avgLength,
            '% Reviews ≤ 30 Days': percentage
        })

        # --- Plots ---
        x = np.arange(len(months))

        # Stacked Bar
        fig1, ax1 = plt.subplots(figsize=(12, 6))
        ax1.bar(x, shortCount, label='≤ 30 Days')
        ax1.bar(x, countExceeds30, bottom=shortCount, label='> 30 Days')
        ax1.set_title('Review Duration Breakdown per Month')
        ax1.set_xticks(x)
        ax1.set_xticklabels(months.keys(), rotation=45)
        ax1.set_ylabel('Number of Projects')
        ax1.legend()
        st.pyplot(fig1)

        # Line Chart
        fig2, ax2 = plt.subplots(figsize=(12, 6))
        ax2.plot(x, avgLength, marker='o')
        ax2.set_title('Average Review Duration per Month')
        ax2.set_xticks(x)
        ax2.set_xticklabels(months.keys(), rotation=45)
        ax2.set_ylabel('Average Days')
        st.pyplot(fig2)

        # --- Export Button ---
        if st.button("📥 Export Excel Report"):
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                excel_path = tmp.name
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
                    workbook = writer.book
                    summary_ws = writer.sheets['Summary']

                    # Format '% Reviews ≤ 30 Days' column as percentage
                    percent_col_idx = summary_df.columns.get_loc('% Reviews ≤ 30 Days') + 1  # 1-based index for Excel
                    for cell in summary_ws.iter_rows(min_row=2, min_col=percent_col_idx, max_col=percent_col_idx):
                        for c in cell:
                            c.number_format = '0.0%'
                        
                    # Format header
                    for cell in summary_ws[1]:
                        cell.font = Font(size=12, bold=True)
                        cell.fill = PatternFill("solid", fgColor="b5dbaf")
                        cell.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)

                    for col in summary_ws.columns:
                            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                            col_letter = col[0].column_letter
                            summary_ws.column_dimensions[col_letter].width = max_length + 5

                    # Charts
                    img_buffer1 = BytesIO()
                    fig1.savefig(img_buffer1, format='png')
                    img_buffer1.seek(0)
                    img1 = XLImage(img_buffer1)
                    summary_ws.add_image(img1, "G1")

                    img_buffer2 = BytesIO()
                    fig2.savefig(img_buffer2, format='png')
                    img_buffer2.seek(0)
                    img2 = XLImage(img_buffer2)
                    summary_ws.add_image(img2, "G31")

                    # Write monthly sheets
                    for name, df in months.items():
                        df.to_excel(writer, sheet_name=name[:31], index=False)
                        ws = writer.sheets[name[:31]]

                        for cell in ws[1]:
                            cell.font = Font(size=12, bold=True)
                            cell.fill = PatternFill("solid", fgColor="b5dbaf")
                            cell.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)

                        for col in ws.columns:
                            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                            col_letter = col[0].column_letter
                            ws.column_dimensions[col_letter].width = max_length + 5

                        # Date formatting
                        for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
                            for cell in row:
                                cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
                        for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
                            for cell in row:
                                cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

                        # Conditional formatting
                        ws.conditional_formatting.add('H2:H1000', CellIsRule(operator = 'greaterThan', formula = ['30'], fill = PatternFill(start_color = 'FFC7CE', end_color = 'FFC7CE', fill_type = 'solid')))

                with open(excel_path, 'rb') as f:
                    st.download_button(
                        label="📂 Download Excel Report",
                        data=f,
                        file_name="monthly_review_data.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
elif section == "🏗️ MDP Annual Report":
    st.header("🏗️ MDP Annual Report")
    st.write("Fetching projects from Asana's portfolio and filtering based on approval date and zoning...")

    # Initialize API
    configuration = asana.Configuration()
    configuration.access_token = st.secrets["ASANA_ACCESS_TOKEN"] 
    api_client = asana.ApiClient(configuration)
    portfolios_api_instance = asana.PortfoliosApi(api_client)

    portfolio_gid = "1205175703519916"
    zoning_port = "1205174022852171"

    opts = {
        'opt_fields': 'custom_fields.name, custom_fields.date_value.date, custom_fields.enum_value.name, custom_fields.number_value, custom_fields.text_value, name',
        'opt_pretty': True
    }

    st.subheader('Filter Options')
    selected_year = st.number_input("Select Year", min_value=2000, max_value=2100, value=2024, step=1)
    start_date = datetime(selected_year, 1, 1).date()
    end_date = datetime(selected_year, 12, 31).date()

    zone_type = st.selectbox('Select Project Type', ['Residential', 'Commercial'])

    if zone_type == 'Residential':
        allowed_projects = {
            'Multi-Family',
            'Residential Single Family Homes',
            'Residential Townhomes',
            'Residential Mixed Density',
            'Residential Duplex or Triplex'
        }

        try:
            api_response = portfolios_api_instance.get_items_for_portfolio(portfolio_gid, opts)
            plat_response = portfolios_api_instance.get_items_for_portfolio(zoning_port, opts)
            projects = list(api_response)
            zones = list(plat_response)

            # ✅ Build plat_lookup ONCE before processing projects
            plat_lookup = {}
            for project in zones:
                if not isinstance(project, dict): continue
                custom_fields = project.get('custom_fields', [])
                project_number = plat_type = None

                for field in custom_fields:
                    if not isinstance(field, dict): continue
                    name = field.get('name')
                    if name == 'Project No':
                        project_number = field.get('text_value')
                    elif name == 'Type of Plat':
                        enum_value = field.get('enum_value')
                        plat_type = enum_value.get('name') if enum_value else None

                if project_number and plat_type:
                    plat_lookup[project_number] = plat_type

            # ✅ Now iterate through main projects
            export_data = []
            SF = TH = Multi = Area = total = matched = 0

            for project in projects:
                if not isinstance(project, dict): continue

                total += 1
                custom_fields = project.get('custom_fields', [])
                approved_date = zoning = project_number = None
                sf_lots = th_lots = mf_units = area_acres = 0

                for field in custom_fields:
                    if not isinstance(field, dict): continue
                    name = field.get('name')
                    if name == 'Date Plan Approved':
                        date_info = field.get('date_value')
                        if date_info and 'date' in date_info:
                            try:
                                approved_date = datetime.strptime(date_info['date'], '%Y-%m-%d').date()
                            except ValueError:
                                continue
                    elif name == 'Proposed Land Use':
                        enum_value = field.get('enum_value')
                        zoning = enum_value.get('name') if enum_value else None
                    elif name == 'Project No':
                        project_number = field.get('text_value')
                    elif name == 'SF Lots':
                        sf_lots = field.get('number_value') or 0
                    elif name == 'TH Lots':
                        th_lots = field.get('number_value') or 0
                    elif name == 'Multi-Family Units':
                        mf_units = field.get('number_value') or 0
                    elif name == 'Total Site Acres':
                        area_acres = field.get('number_value') or 0

                if not approved_date or not (start_date <= approved_date <= end_date):
                    continue
                if zoning not in allowed_projects:
                    continue

                matched += 1
                SF += sf_lots
                TH += th_lots
                Multi += mf_units
                Area += area_acres
                total_units = sf_lots + th_lots + mf_units

                plat_type = plat_lookup.get(project_number)

                export_data.append({
                    'Project Name': project.get('name'),
                    'Project Number': project_number,
                    'Approval Date': approved_date,
                    'Zoning': zoning,
                    'SF Lots': sf_lots,
                    'TH Lots': th_lots,
                    'Multi-Family Units': mf_units,
                    'Total Units/Lots': total_units,
                    'Area (Acres)': area_acres,
                    'Plat': plat_type
                })

            df = pd.DataFrame(export_data)
            df = pd.concat([df, pd.DataFrame([{
                'Project Name': 'TOTAL',
                'SF Lots': SF,
               'TH Lots': TH,
               'Multi-Family Units': Multi,
                'Total Units/Lots': SF + TH + Multi,
                'Area (Acres)': Area,
                'Plat': ''  # total row has no single plat type
            }])], ignore_index=True)

            st.subheader("📋 Summary")
            st.write(f"Total projects in portfolio: **{total}**")
            st.write(f"Projects approved in **{selected_year}** with allowed land use: **{matched}**")
            st.write(f"SF Lots: **{SF}**, TH Lots: **{TH}**, Multi-Family Units: **{Multi}**, Total Units: **{SF + TH + Multi}**, Area: **{Area} acres**")
            st.write(plat_lookup)

            st.dataframe(df)

            output = BytesIO()
            df.to_excel(output, index=False)
            st.download_button(
                label="📥 Download Excel File",
                data=output.getvalue(),
                file_name="approved_projects_2024.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except ApiException as e:
            st.error(f"API Exception: {e}")


    elif zone_type == 'Commercial': 
        allowed_projects = {
            'Commercial'
        }
        try:
            api_response = portfolios_api_instance.get_items_for_portfolio(portfolio_gid, opts)
            projects = list(api_response)
        
            export_data = []
            CommercialSQ = Area = total = matched = 0

            for project in projects:
                if not isinstance(project, dict): continue

                total += 1
                custom_fields = project.get('custom_fields', [])
                approved_date = zoning = project_number = None
                sqft = area_acres = 0

                for field in custom_fields:
                    if not isinstance(field, dict): continue
                    name = field.get('name')
                    if name == 'Date Plan Approved':
                        date_info = field.get('date_value')
                        if date_info and 'date' in date_info:
                            try:
                                approved_date = datetime.strptime(date_info['date'], '%Y-%m-%d').date()
                            except ValueError:
                                continue
                    elif name == 'Proposed Land Use':
                        enum_value = field.get('enum_value')
                        zoning = enum_value.get('name') if enum_value else None
                    elif name == 'Project No':
                        project_number = field.get('text_value')
                    elif name == 'Commercial Square Feet':
                        sqft = field.get('number_value') or 0
                    elif name == 'Total Site Acres':
                        area_acres = field.get('number_value') or 0

                if not approved_date or not (start_date <= approved_date <= end_date):
                    continue
                if zoning not in allowed_projects:
                    continue

                matched += 1
                CommercialSQ += sqft
                Area += area_acres

                export_data.append({
                    'Project Name': project.get('name'),
                    'Project Number': project_number,
                    'Approval Date': approved_date,
                    'Zoning': zoning,
                    'Commercial Square Feet': sqft,
                    'Area (Acres)': area_acres,
                    'Plat': resub
                })

            df = pd.DataFrame(export_data)
            df = pd.concat([
                df,
                pd.DataFrame([{
                    'Project Name': 'TOTAL',
                    'Commercial Square Feet': CommercialSQ,
                    'Area (Acres)': Area
                }])
            ], ignore_index=True)
            st.subheader("📋 Summary")
            st.write(f"Total projects in portfolio: **{total}**")
            st.write(f"Projects approved in 2024 with allowed land use: **{matched}**")
            st.write(f"Commercial Square Feet **{CommercialSQ}**, Area: **{Area} acres**")

            st.dataframe(df)

            # Excel Export
            output = BytesIO()
            df.to_excel(output, index=False)
            st.download_button(
                label="📥 Download Excel File",
                data=output.getvalue(),
                file_name="approved_projects_2024.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except ApiException as e:
            st.error(f"API Exception: {e}")